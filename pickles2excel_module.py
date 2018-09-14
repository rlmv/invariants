import pyphi
import pickle
import os
import numpy as np
import string
import openpyxl
from openpyxl import Workbook
import pandas as pd
from pandas import DataFrame, Series
from collections import OrderedDict
import glob

def pickles2excel(network_pickle, state):
    # Create an excel from a list of pickles in the same location of the script that calls the function
    # Args:
    # network_pickle: path/network_filename.pickle
    # state: tuple of ints with state of each element of the network (e.g. np.zeros(20))
    # To use the function type in script: from pickles2excel_module import *
    # Put pickles2excel_module.py in the same folder of the script
    
    network = pd.read_pickle(open(network_pickle, 'rb'))
    state = state
    subsystem = pyphi.Subsystem(network, state, range(network.size))

    excel_path = os.path.basename(network_pickle).replace("_network.pickle", ".xlsx")
    
    wb = Workbook()
    wb.save(excel_path)
    
    book = openpyxl.load_workbook(excel_path)
    writer = pd.ExcelWriter(excel_path, engine='openpyxl')
    writer.book = book
    sheet = book.worksheets[0]
    startrow = sheet.max_row

    pickles = sorted(sorted(glob.iglob('*.pickle')),key=len)
    
    for pickle in pickles:
        
        concept = pd.read_pickle(pickle)
        
        if concept.phi:
            pandas_series = Series(OrderedDict({
                'mechanism':  subsystem.indices2nodes(concept.mechanism),
                'cause_purview': subsystem.indices2nodes(concept.cause_purview),
                'effect_purview': subsystem.indices2nodes(concept.effect_purview),
                'cause_phi': concept.cause.phi,
                'effect_phi': concept.effect.phi,
            }))
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            row = DataFrame(pandas_series).transpose()
            row.to_excel(writer, sheet_name='Sheet', startrow=startrow, index=False, header=None)
            startrow += 1
            writer.save()
    
    print(f'Success! Pickles saved to {excel_path}')
